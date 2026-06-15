from __future__ import annotations

import argparse
import json
from pathlib import Path
from unicodedata import normalize

from openpyxl import load_workbook


CHANNEL_COLUMNS = [
    "Caracol TV",
    "RCN Televisión",
    "DSports",
    "Win+",
    "Paramount+",
    "Disney+",
]

TEAM_ALIASES = {
    "ALEMANIA": "Germany",
    "ARABIA SAUDITA": "Saudi Arabia",
    "ARGELIA": "Algeria",
    "ARGENTINA": "Argentina",
    "AUSTRALIA": "Australia",
    "AUSTRIA": "Austria",
    "BELGICA": "Belgium",
    "BOSNIA": "Bosnia and Herzegovina",
    "BRASIL": "Brazil",
    "CABO VERDE": "Cabo Verde",
    "CANADA": "Canada",
    "COLOMBIA": "Colombia",
    "COREA DEL SUR": "South Korea",
    "COSTA DE MARFIL": "Côte d'Ivoire",
    "CROACIA": "Croatia",
    "CURAZAO": "Curaçao",
    "ECUADOR": "Ecuador",
    "EGIPTO": "Egypt",
    "ESCOCIA": "Scotland",
    "ESPANA": "Spain",
    "ESTADOS UNIDOS": "United States",
    "FRANCIA": "France",
    "GHANA": "Ghana",
    "HAITI": "Haiti",
    "INGLATERRA": "England",
    "IRAK": "Iraq",
    "IRAN": "IR Iran",
    "JAPON": "Japan",
    "JORDANIA": "Jordan",
    "MARRUECOS": "Morocco",
    "MEXICO": "Mexico",
    "NORUEGA": "Norway",
    "NUEVA ZELANDA": "New Zealand",
    "PAISES BAJOS": "Netherlands",
    "PANAMA": "Panama",
    "PARAGUAY": "Paraguay",
    "PORTUGAL": "Portugal",
    "QATAR": "Qatar",
    "RD CONGO": "DR Congo",
    "REP CHECA": "Czech Republic",
    "SENEGAL": "Senegal",
    "SUDÁFRICA": "South Africa",
    "SUDAFRICA": "South Africa",
    "SUECIA": "Sweden",
    "SUIZA": "Switzerland",
    "TUNEZ": "Tunisia",
    "TURQUIA": "Türkiye",
    "URUGUAY": "Uruguay",
    "UZBEKISTAN": "Uzbekistan",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Build config/televisacion.json from the broadcast schedule Excel.")
    parser.add_argument("excel_path", type=Path)
    parser.add_argument("--output", type=Path, default=Path("config/televisacion.json"))
    parser.add_argument("--schedule", type=Path, default=Path("config/calendario_partidos.json"))
    args = parser.parse_args()

    workbook = load_workbook(args.excel_path, data_only=True)
    sheet = workbook["Partidos"]
    headers = [cell.value for cell in sheet[1]]
    indexes = {header: idx for idx, header in enumerate(headers)}

    missing = [column for column in ["No.", *CHANNEL_COLUMNS] if column not in indexes]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    schedule = json.loads(args.schedule.read_text(encoding="utf-8"))["matches"]
    schedule_by_teams = {
        frozenset([_canonical_team(match["team_a"]), _canonical_team(match["team_b"])]): match["match_id"]
        for match in schedule
    }

    matches: dict[str, list[str]] = {}
    remapped: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        match_number = row[indexes["No."]]
        if not match_number:
            continue
        excel_match_id = f"M{int(match_number):03d}"
        team_key = frozenset([
            _canonical_team(_translate_team(row[indexes["Equipo 1"]])),
            _canonical_team(_translate_team(row[indexes["Equipo 2"]])),
        ])
        match_id = schedule_by_teams.get(team_key)
        if not match_id:
            raise ValueError(f"Could not match Excel row {excel_match_id}: {row[indexes['Partido']]}")
        if match_id != excel_match_id:
            remapped.append(f"{excel_match_id}->{match_id}")
        matches[match_id] = [
            channel
            for channel in CHANNEL_COLUMNS
            if row[indexes[channel]] == "X"
        ]

    if len(matches) != len(schedule):
        raise ValueError(f"Expected {len(schedule)} matches, got {len(matches)}")

    payload = {
        "source": {
            "name": "Calendario Mundial 2026 - canales por partido",
            "url": "https://www.jorgebarajas.com/wp-content/uploads/2026/04/Calendario-Copa-Mundial-de-la-FIFA-2026-Tabloide-4.pdf",
            "validated_with": args.excel_path.name,
            "retrieved_at": "2026-06-14",
        },
        "channels": CHANNEL_COLUMNS,
        "matches": matches,
    }
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(matches)} matches to {args.output}")
    if remapped:
        print("Remapped by team matching: " + ", ".join(remapped))
    return 0


def _translate_team(value: object) -> str:
    raw = str(value or "").strip()
    key = _normalize_text(raw)
    return TEAM_ALIASES.get(key, raw)


def _canonical_team(value: object) -> str:
    return _normalize_text(str(value or ""))


def _normalize_text(value: str) -> str:
    decomposed = normalize("NFKD", value)
    ascii_text = "".join(char for char in decomposed if not char.encode("ascii", "ignore") == b"")
    cleaned = "".join(char if char.isalnum() else " " for char in ascii_text.upper())
    return " ".join(cleaned.split())


if __name__ == "__main__":
    raise SystemExit(main())
