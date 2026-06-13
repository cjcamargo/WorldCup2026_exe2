from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from polla.config import ROOT, config_path, data_path, load_json, output_path
from polla.results import load_confirmed_results
from polla.schedule import load_schedule
from polla.store import GoogleSheetsStore, hash_pin


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inicializa el Google Sheet backend de la app Streamlit.")
    parser.add_argument("--spreadsheet-id", help="ID de un Google Sheet existente. Si se omite, se crea uno nuevo.")
    parser.add_argument("--title", default="Polla Mundialista App DB", help="Titulo si se crea un Google Sheet nuevo.")
    parser.add_argument("--service-account-json", help="Ruta al JSON de service account. Alternativa: GOOGLE_SERVICE_ACCOUNT_JSON.")
    parser.add_argument("--default-pin", default="1234", help="PIN inicial para participantes.")
    parser.add_argument("--admin-pin", default="admin123", help="PIN inicial para usuario admin.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    credentials = _load_credentials(args.service_account_json)
    spreadsheet_id = args.spreadsheet_id or _create_spreadsheet(credentials, args.title)
    store = GoogleSheetsStore(spreadsheet_id, credentials)
    store.ensure_schema()

    participants = load_json(config_path("participantes.json"))["participants"]
    users = [
        {
            "participant": item["name"],
            "pin_hash": hash_pin(item["name"], args.default_pin),
            "role": "player",
            "active": str(bool(item.get("active", True))),
        }
        for item in participants
    ]
    users.append({
        "participant": "admin",
        "pin_hash": hash_pin("admin", args.admin_pin),
        "role": "admin",
        "active": "True",
    })
    store.replace_rows("Users", users)
    store.replace_rows("Matches", [_match_row(match) for match in load_schedule(config_path("calendario_partidos.json"))])
    store.replace_rows("Results", [_result_row(result) for result in _dedupe_results(load_confirmed_results(data_path("resultados", "resultados_confirmados.csv")))])
    store.replace_rows("Settings", _initial_settings())
    ranking_rows = _ranking_rows(output_path("ranking_polla.xlsx"))
    if ranking_rows:
        store.replace_rows("Ranking", ranking_rows)
    print("Google Sheet backend listo:")
    print(f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit")
    print(f"PIN inicial participantes: {args.default_pin}")
    print(f"PIN inicial admin: {args.admin_pin}")
    return 0


def _load_credentials(path: str | None) -> dict:
    if path:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    raw = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if raw:
        return json.loads(raw)
    raise RuntimeError("Define --service-account-json o GOOGLE_SERVICE_ACCOUNT_JSON.")


def _create_spreadsheet(credentials: dict, title: str) -> str:
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    client = gspread.authorize(Credentials.from_service_account_info(credentials, scopes=scopes))
    spreadsheet = client.create(title)
    return spreadsheet.id


def _match_row(match) -> dict:
    return {
        "match_id": match.match_id,
        "phase": match.phase,
        "team_a": match.team_a,
        "team_b": match.team_b,
        "kickoff_at": match.kickoff_at.isoformat() if match.kickoff_at else "",
        "status": match.status,
    }


def _result_row(result) -> dict:
    return {
        "match_id": result.match_id,
        "team_a": result.team_a,
        "team_b": result.team_b,
        "goals_a_real": result.goals_a_real,
        "goals_b_real": result.goals_b_real,
        "status": result.status,
        "phase": result.phase,
        "kickoff_at": result.kickoff_at.isoformat() if result.kickoff_at else "",
        "source": result.source,
        "source_url": result.source_url,
        "confirmed": str(result.confirmed),
    }


def _dedupe_results(results) -> list:
    by_id = {}
    for result in results:
        by_id[result.match_id] = result
    return list(by_id.values())


def _initial_settings() -> list[dict]:
    rows = [
        {"key": "final_picks_closed", "value": "False"},
        {"key": "actual_champion", "value": ""},
        {"key": "actual_runner_up", "value": ""},
        {"key": "actual_third_place", "value": ""},
    ]
    for group in [f"Group {chr(code)}" for code in range(ord("A"), ord("L") + 1)]:
        rows.append({"key": f"group_closed_{group}", "value": "False"})
    return rows


def _ranking_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Ranking"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


if __name__ == "__main__":
    raise SystemExit(main())
