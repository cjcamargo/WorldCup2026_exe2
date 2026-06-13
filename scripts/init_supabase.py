from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from polla.config import config_path, data_path, load_json, output_path
from polla.results import load_confirmed_results
from polla.schedule import load_schedule
from polla.store import hash_pin
from polla.supabase_store import SupabaseStore


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inicializa tablas Supabase de la app Streamlit.")
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL"), help="Supabase project URL.")
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"), help="Supabase service role key.")
    parser.add_argument("--default-pin", default="1234", help="PIN inicial para participantes.")
    parser.add_argument("--admin-pin", default="admin123", help="PIN inicial para usuario admin.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.url or not args.service_role_key:
        raise RuntimeError("Define --url/--service-role-key o SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY.")
    store = SupabaseStore(args.url, args.service_role_key)

    participants = load_json(config_path("participantes.json"))["participants"]
    users = [
        {
            "participant": item["name"],
            "pin_hash": hash_pin(item["name"], args.default_pin),
            "role": "player",
            "active": bool(item.get("active", True)),
        }
        for item in participants
    ]
    users.append({
        "participant": "admin",
        "pin_hash": hash_pin("admin", args.admin_pin),
        "role": "admin",
        "active": True,
    })
    store.replace_rows("users", users)
    store.replace_rows("matches", [_match_row(match) for match in load_schedule(config_path("calendario_partidos.json"))])
    store.replace_rows("results", [_result_row(result) for result in _dedupe_results(load_confirmed_results(data_path("resultados", "resultados_confirmados.csv")))])
    store.replace_rows("settings", _initial_settings())
    ranking_rows = _ranking_rows(output_path("ranking_polla.xlsx"))
    if ranking_rows:
        store.replace_rows("ranking", ranking_rows)

    print("Supabase backend listo.")
    print(f"PIN inicial participantes: {args.default_pin}")
    print(f"PIN inicial admin: {args.admin_pin}")
    return 0


def _match_row(match) -> dict:
    return {
        "match_id": match.match_id,
        "phase": match.phase,
        "team_a": match.team_a,
        "team_b": match.team_b,
        "kickoff_at": match.kickoff_at.isoformat() if match.kickoff_at else None,
        "status": match.status,
    }


def _result_row(result) -> dict:
    return {
        "match_id": result.match_id,
        "team_a": result.team_a,
        "team_b": result.team_b,
        "goals_a_real": result.goals_a_real,
        "goals_b_real": result.goals_b_real,
        "status": result.status,
        "phase": result.phase,
        "kickoff_at": result.kickoff_at.isoformat() if result.kickoff_at else None,
        "source": result.source,
        "source_url": result.source_url,
        "confirmed": result.confirmed,
    }


def _dedupe_results(results) -> list:
    by_id = {}
    for result in results:
        by_id[result.match_id] = result
    return list(by_id.values())


def _initial_settings() -> list[dict]:
    rows = [
        {"key": "final_picks_closed", "value": "False"},
        {"key": "actual_champion", "value": ""},
        {"key": "actual_runner_up", "value": ""},
        {"key": "actual_third_place", "value": ""},
    ]
    for group in [f"Group {chr(code)}" for code in range(ord("A"), ord("L") + 1)]:
        rows.append({"key": f"group_closed_{group}", "value": "False"})
    return rows


def _ranking_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Ranking"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


if __name__ == "__main__":
    raise SystemExit(main())
