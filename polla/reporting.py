from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_workbook(path: Path, ranking: list[dict[str, Any]], detail: list[dict[str, Any]], warnings: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"
    _write_rows(ws, ranking or [{"participant": "", "points": 0, "rank": ""}])
    detail_ws = wb.create_sheet("Detalle")
    _write_rows(detail_ws, detail or [{"participant": "", "match_id": "", "points": 0}])
    warn_ws = wb.create_sheet("Alertas")
    _write_rows(warn_ws, [{"warning": warning} for warning in warnings] or [{"warning": ""}])
    for sheet in wb.worksheets:
        _format_sheet(sheet)
    wb.save(path)


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 45)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width
