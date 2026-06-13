from __future__ import annotations

import hashlib
import re
from dataclasses import replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import FinalPicks, Prediction


SCORE_RE = re.compile(r"^\s*(\d{1,2})\s*[-:]\s*(\d{1,2})\s*$")


def normalize_team(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return None


def winner(team_a: str, team_b: str, goals_a: int | None, goals_b: int | None) -> str | None:
    if goals_a is None or goals_b is None:
        return None
    if goals_a > goals_b:
        return team_a
    if goals_b > goals_a:
        return team_b
    return "Empate"


def workbook_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_predictions(path: Path, participant: str, extractor_config: dict) -> tuple[list[Prediction], FinalPicks, list[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    mode = extractor_config.get("mode", "heuristic")
    if mode == "manual":
        predictions = _read_manual(wb, participant, extractor_config)
    elif mode == "ingreso":
        predictions = _read_ingreso(wb, participant, extractor_config)
    else:
        predictions = _read_heuristic(wb, participant, extractor_config)
    picks = _read_final_picks(wb, participant, extractor_config)
    warnings: list[str] = []
    if not predictions:
        warnings.append(f"No se detectaron pronosticos para {participant}; revisar config/extractor.json.")
    wb.close()
    return predictions, picks, warnings


def apply_prediction_overrides(
    predictions: list[Prediction],
    overrides: list[dict[str, Any]] | None,
) -> tuple[list[Prediction], list[str]]:
    if not overrides:
        return predictions, []
    by_key = {
        (str(item.get("participant", "")).strip(), str(item.get("match_id", "")).strip()): item
        for item in overrides
        if item.get("participant") and item.get("match_id")
    }
    if not by_key:
        return predictions, []
    updated: list[Prediction] = []
    warnings: list[str] = []
    for pred in predictions:
        override = by_key.get((pred.participant, pred.match_id))
        if not override:
            updated.append(pred)
            continue
        goals_a = to_int(override.get("goals_a_pred"))
        goals_b = to_int(override.get("goals_b_pred"))
        updated.append(replace(
            pred,
            goals_a_pred=goals_a,
            goals_b_pred=goals_b,
            winner_pred=winner(pred.team_a, pred.team_b, goals_a, goals_b),
            valid=True,
            invalid_reason=None,
        ))
        warnings.append(
            f"Override manual aplicado a {pred.participant} {pred.match_id}: {goals_a}-{goals_b}"
        )
    return updated, warnings


def _read_ingreso(wb, participant: str, config: dict) -> list[Prediction]:
    ingreso_cfg = config.get("ingreso", {})
    sheet_name = ingreso_cfg.get("sheet", "Ingreso")
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    match_col = ingreso_cfg.get("match_id_col", "B")
    team_a_col = ingreso_cfg.get("team_a_col", "F")
    goals_a_col = ingreso_cfg.get("goals_a_col", "G")
    goals_b_col = ingreso_cfg.get("goals_b_col", "J")
    team_b_col = ingreso_cfg.get("team_b_col", "L")
    first_header_row = int(ingreso_cfg.get("first_header_row", 8))
    row_step = int(ingreso_cfg.get("row_step", 6))
    last_header_row = int(ingreso_cfg.get("last_header_row") or ws.max_row)

    predictions: list[Prediction] = []
    for header_row in range(first_header_row, last_header_row + 1, row_step):
        match_number = to_int(ws[f"{match_col}{header_row}"].value)
        if match_number is None:
            continue
        score_row = header_row + 1
        team_a = normalize_team(ws[f"{team_a_col}{score_row}"].value)
        team_b = normalize_team(ws[f"{team_b_col}{score_row}"].value)
        if not team_a or not team_b:
            continue
        goals_a = to_int(ws[f"{goals_a_col}{score_row}"].value)
        goals_b = to_int(ws[f"{goals_b_col}{score_row}"].value)
        predictions.append(Prediction(
            participant=participant,
            match_id=f"M{match_number:03d}",
            phase=ingreso_cfg.get("phase", "Group stage"),
            team_a=team_a,
            team_b=team_b,
            goals_a_pred=goals_a,
            goals_b_pred=goals_b,
            winner_pred=winner(team_a, team_b, goals_a, goals_b),
            source_sheet=sheet_name,
            source_row=score_row,
        ))
    return predictions


def _read_manual(wb, participant: str, config: dict) -> list[Prediction]:
    predictions: list[Prediction] = []
    for item in config.get("manual_match_ranges", []):
        ws = wb[item["sheet"]]
        team_a = normalize_team(ws[item["team_a_cell"]].value)
        team_b = normalize_team(ws[item["team_b_cell"]].value)
        goals_a = to_int(ws[item["goals_a_cell"]].value)
        goals_b = to_int(ws[item["goals_b_cell"]].value)
        if not team_a or not team_b:
            continue
        predictions.append(Prediction(
            participant=participant,
            match_id=item["match_id"],
            phase=item.get("phase"),
            team_a=team_a,
            team_b=team_b,
            goals_a_pred=goals_a,
            goals_b_pred=goals_b,
            winner_pred=winner(team_a, team_b, goals_a, goals_b),
            source_sheet=item["sheet"],
        ))
    return predictions


def _read_heuristic(wb, participant: str, config: dict) -> list[Prediction]:
    include = set(config.get("sheets", {}).get("include") or [])
    exclude = set(config.get("sheets", {}).get("exclude") or [])
    predictions: list[Prediction] = []
    seen: set[tuple[str, str, int, int, str]] = set()
    for ws in wb.worksheets:
        if include and ws.title not in include:
            continue
        if ws.title in exclude:
            continue
        for row_number, row in enumerate(ws.iter_rows(), start=1):
            values = [cell.value for cell in row]
            row_predictions = _detect_row(values)
            for team_a, team_b, goals_a, goals_b in row_predictions:
                key = (team_a, team_b, goals_a, goals_b, ws.title)
                if key in seen:
                    continue
                seen.add(key)
                match_id = f"{_slug(team_a)}_vs_{_slug(team_b)}"
                predictions.append(Prediction(
                    participant=participant,
                    match_id=match_id,
                    team_a=team_a,
                    team_b=team_b,
                    goals_a_pred=goals_a,
                    goals_b_pred=goals_b,
                    winner_pred=winner(team_a, team_b, goals_a, goals_b),
                    source_sheet=ws.title,
                    source_row=row_number,
                ))
    return predictions


def _detect_row(values: list[Any]) -> list[tuple[str, str, int, int]]:
    out: list[tuple[str, str, int, int]] = []
    cleaned = [normalize_team(v) for v in values]
    ints = [(idx, to_int(v)) for idx, v in enumerate(values)]
    int_positions = [(idx, val) for idx, val in ints if val is not None and 0 <= val <= 20]
    for idx, text in enumerate(cleaned):
        if not text:
            continue
        match = SCORE_RE.match(text)
        if match:
            left_team = _nearest_text(cleaned, idx, -1)
            right_team = _nearest_text(cleaned, idx, 1)
            if left_team and right_team:
                out.append((left_team, right_team, int(match.group(1)), int(match.group(2))))
    for pos_a, goals_a in int_positions:
        for pos_b, goals_b in int_positions:
            if pos_b <= pos_a or pos_b - pos_a > 4:
                continue
            left_team = _nearest_text(cleaned, pos_a, -1)
            right_team = _nearest_text(cleaned, pos_b, 1)
            if left_team and right_team and left_team != right_team:
                out.append((left_team, right_team, goals_a, goals_b))
    return out


def _nearest_text(values: list[str | None], start: int, step: int) -> str | None:
    idx = start + step
    while 0 <= idx < len(values):
        value = values[idx]
        if value and not SCORE_RE.match(value) and not value.isdigit() and len(value) > 2:
            return value
        idx += step
    return None


def _read_final_picks(wb, participant: str, config: dict) -> FinalPicks:
    picks_cfg = config.get("final_picks", {})
    sheet_name = picks_cfg.get("sheet")
    if not sheet_name:
        return FinalPicks(participant=participant)
    ws = wb[sheet_name]
    return FinalPicks(
        participant=participant,
        champion=normalize_team(ws[picks_cfg["champion_cell"]].value) if picks_cfg.get("champion_cell") else None,
        runner_up=normalize_team(ws[picks_cfg["runner_up_cell"]].value) if picks_cfg.get("runner_up_cell") else None,
        third_place=normalize_team(ws[picks_cfg["third_place_cell"]].value) if picks_cfg.get("third_place_cell") else None,
    )


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
